import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 输入路径
file1 = r"D:\PythonProject2\holo_new2_3level_100-100\output\origin_gray_levels_dominant.xlsx"
file2 = r"D:\PythonProject2\holo_new2_3level_100-100\output\print_gray_levels_dominant.xlsx"

# 输出路径
output_file = r"D:\PythonProject2\holo_new2_3level_100-100\output\compare\comparison_result.xlsx"
stats_file = r"D:\PythonProject2\holo_new2_3level_100-100\output\compare\comparison_stats.xlsx"

# 读取 Excel 数据
df1 = pd.read_excel(file1, header=None)
df2 = pd.read_excel(file2, header=None)

# 确保尺寸一致
if df1.shape != df2.shape:
    raise ValueError(f"两个Excel尺寸不一致: {df1.shape} vs {df2.shape}")

# 执行逐元素比较
comparison_result = df1.values == df2.values  # numpy array of bool

# 统计TRUE和FALSE的数量
true_count = np.sum(comparison_result)
false_count = comparison_result.size - true_count

# 创建统计结果的DataFrame
stats_df = pd.DataFrame({
    'Comparison Type': ['TRUE (Matches)', 'FALSE (Mismatches)'],
    'Count': [true_count, false_count]
})


# 保存比较结果（带颜色）
def save_colored_comparison(df, output_path):
    # 创建一个新的工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 将DataFrame写入工作表
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # 应用颜色
            if value:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色
            else:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色

    # 保存工作簿
    wb.save(output_path)


# 保存比较结果（带颜色）
save_colored_comparison(pd.DataFrame(comparison_result), output_file)

# 保存统计结果
stats_df.to_excel(stats_file, index=False)

print(f"比较结果已保存至: {output_file} (TRUE为绿色，FALSE为红色)")
print(f"统计结果已保存至: {stats_file}")